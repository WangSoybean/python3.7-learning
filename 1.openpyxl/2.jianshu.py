from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
#新建文件
wb = Workbook()
#激活表单
ws = wb.active
#1.表单操作
#建立另一个表单
ws1 = wb.create_sheet("My_sheet")
#根据表单位置建立表单
ws2 = wb.create_sheet("my sheet2", 1)
ws3 = wb.create_sheet("my sheet3", -1)
#更改表单名字
ws1.title = 'New sheet'
#删除表单
del wb["New sheet"]
#列出所有表单的两种方式
#方式1：
print(wb.sheetnames)
#方式2：
for sheet in wb:
    print(sheet.title)

#2.写入表单数据
ws['A1'] = 'Class'
ws['B1'] = 'Price'
ws['C1'] = 'Amount'
ws.append(["Apple", "8", "5.5"])
ws.append(["Orange", "6", "4.5"])
ws.append(["Grape", "12", "6.5"])
wb.save("test.xlsx")

#3.读取表单
wb_test = load_workbook("test.xlsx")
#打印表单
print(wb_test.sheetnames)
#选取表单
Sheet = wb_test['Sheet']
#表单行数、列数
print("该表格共{}行".format(Sheet.max_row))
print("该表格共{}列".format(Sheet.max_column))
#输出某一单元格的值
C = Sheet['C1']
print(C.value)
#输出表格前4行的内容
for row in Sheet.iter_rows(min_row=1, max_row=4, values_only=True):
    print(row)
#计算总价
for row in range(2, Sheet.max_row + 1):
    price = Sheet["B{}".format(row)].value
    amount = Sheet["C{}".format(row)].value
    Sheet["D{}".format(row)] = float(price) * float(amount)
Sheet["D1"] = "Total"
wb_test.save("test1.xlsx")

#4.插入图片
img = Image('image.png')
Sheet.add_image(img, 'E1')
wb_test.save("test1.xlsx")


